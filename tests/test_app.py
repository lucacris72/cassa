from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

from sqlalchemy import select

from restaurant_pos.app.auth import hash_pin
from restaurant_pos.app.models import Category, Order, OrderItem, Printer, PrintJob, Product, ProductImportAlias, RegisterClosure, Reservation, ReservationItem, User
from restaurant_pos.app.services import printing
from restaurant_pos.app.services.numbering import business_date_for
from restaurant_pos.app.services.orders import create_confirmed_order, parse_cart_json
from restaurant_pos.app.services.reservations import import_reservations_from_xlsx


def test_seed_data(db_session):
    assert db_session.scalar(select(User).where(User.name == "admin")) is not None
    assert db_session.scalar(select(User).where(User.name == "cashier")) is not None
    assert db_session.scalar(select(Printer).where(Printer.is_customer_printer.is_(True))) is not None
    assert db_session.scalar(select(Category).where(Category.name == "Cucina")) is not None
    assert db_session.scalar(select(Product).where(Product.name == "Panino salamella")) is not None


def test_login_valid_and_invalid(client):
    page = client.get("/login")
    assert page.status_code == 200
    assert "Cassa" in page.text

    invalid = client.post("/login", data={"pin": "9999"}, follow_redirects=False)
    assert invalid.status_code == 303
    assert invalid.headers["location"] == "/login"

    valid = client.post("/login", data={"pin": "1234"}, follow_redirects=False)
    assert valid.status_code == 303
    assert valid.headers["location"] == "/"


def test_main_pages_render(admin_client):
    for path in ["/", "/orders", "/reservations", "/closures", "/products", "/categories", "/printers", "/users", "/mobile"]:
        response = admin_client.get(path)
        assert response.status_code == 200, path


def test_healthcheck(client):
    response = client.get("/healthz")
    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def test_admin_crud_pages(admin_client, db_session):
    printer_response = admin_client.post(
        "/printers",
        data={"name": "Expo", "type": "fake", "port": "9100", "enabled": "true", "partial_cut": "true"},
        follow_redirects=False,
    )
    assert printer_response.status_code == 303
    printer = db_session.scalar(select(Printer).where(Printer.name == "Expo"))
    assert printer is not None
    assert printer.partial_cut is True
    printers_page = admin_client.get("/printers")
    assert printers_page.status_code == 200
    assert "Parziale" in printers_page.text

    category_response = admin_client.post(
        "/categories",
        data={"name": "Dolci", "printer_id": str(printer.id), "sort_order": "40"},
        follow_redirects=False,
    )
    assert category_response.status_code == 303
    category = db_session.scalar(select(Category).where(Category.name == "Dolci"))
    assert category is not None

    product_response = admin_client.post(
        "/products",
        data={"name": "Torta", "price": "4.50", "category_id": str(category.id), "sort_order": "10"},
        follow_redirects=False,
    )
    assert product_response.status_code == 303
    product = db_session.scalar(select(Product).where(Product.name == "Torta"))
    assert product is not None
    assert product.price_cents == 450

    products_page = admin_client.get(f"/products?category_id={category.id}&show=all")
    assert products_page.status_code == 200
    assert "Torta" in products_page.text
    assert "Solo attivi" in products_page.text

    cucina = db_session.scalar(select(Category).where(Category.name == "Cucina"))
    edit_response = admin_client.post(
        f"/products/{product.id}/edit",
        data={
            "name": "Torta aggiornata",
            "price": "5.00",
            "category_id": str(cucina.id),
            "sort_order": "3",
            "active": "false",
            "description": "fetta",
            "return_to": "/products?show=all",
        },
        follow_redirects=False,
    )
    assert edit_response.status_code == 303
    assert edit_response.headers["location"] == "/products?show=all"
    db_session.refresh(product)
    assert product.name == "Torta aggiornata"
    assert product.price_cents == 500
    assert product.category_id == cucina.id
    assert product.sort_order == 3
    assert product.active is False

    toggle_response = admin_client.post(
        f"/products/{product.id}/toggle-active",
        data={"return_to": "/products?show=inactive"},
        follow_redirects=False,
    )
    assert toggle_response.status_code == 303
    assert toggle_response.headers["location"] == "/products?show=inactive"
    db_session.refresh(product)
    assert product.active is True


def test_category_can_be_hidden_from_cashier_without_deactivation(admin_client, db_session):
    category = db_session.scalar(select(Category).where(Category.name == "Cucina"))
    assert category is not None

    response = admin_client.post(f"/categories/{category.id}/toggle-cashier-visibility", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/categories"

    db_session.expire_all()
    category = db_session.get(Category, category.id)
    assert category.active is True
    assert category.show_in_cashier is False

    categories_page = admin_client.get("/categories")
    assert categories_page.status_code == 200
    assert "Nascosta" in categories_page.text

    cashier_page = admin_client.get("/")
    assert cashier_page.status_code == 200
    assert f'data-category-filter="{category.id}"' not in cashier_page.text
    assert "Panino salamella" not in cashier_page.text

    mobile_page = admin_client.get("/mobile")
    assert mobile_page.status_code == 200
    assert f'data-category-filter="{category.id}"' not in mobile_page.text
    assert "Panino salamella" not in mobile_page.text


def test_category_delete_requires_no_products(admin_client, db_session):
    response = admin_client.post(
        "/categories",
        data={"name": "Categoria vuota", "printer_id": "0", "sort_order": "99"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    empty_category = db_session.scalar(select(Category).where(Category.name == "Categoria vuota"))
    assert empty_category is not None
    empty_category_id = empty_category.id

    categories_page = admin_client.get("/categories")
    assert categories_page.status_code == 200
    assert "Elimina" in categories_page.text

    delete_empty = admin_client.post(f"/categories/{empty_category_id}/delete", follow_redirects=False)
    assert delete_empty.status_code == 303
    assert delete_empty.headers["location"] == "/categories"
    db_session.expire_all()
    assert db_session.get(Category, empty_category_id) is None

    cucina = db_session.scalar(select(Category).where(Category.name == "Cucina"))
    assert cucina is not None
    cucina_id = cucina.id
    product_count = len(cucina.products)
    assert product_count > 0

    delete_used = admin_client.post(f"/categories/{cucina_id}/delete", follow_redirects=False)
    assert delete_used.status_code == 303
    db_session.expire_all()
    assert db_session.get(Category, cucina_id) is not None
    assert len(db_session.get(Category, cucina_id).products) == product_count


def test_long_product_names_have_stable_cashier_markup(admin_client, db_session):
    category = db_session.scalar(select(Category).where(Category.name == "Cucina"))
    long_name = "INSALATA MISTA - SENZA GLUTINE E SENZA LATTOSIO"
    db_session.add(Product(name=long_name, price_cents=700, category_id=category.id, active=True, sort_order=999))
    db_session.commit()

    cashier_page = admin_client.get("/")
    assert cashier_page.status_code == 200
    assert f'title="{long_name}"' in cashier_page.text
    assert '<span class="product-name">INSALATA MISTA - SENZA GLUTINE E SENZA LATTOSIO</span>' in cashier_page.text

    mobile_page = admin_client.get("/mobile")
    assert mobile_page.status_code == 200
    assert f'title="{long_name}"' in mobile_page.text
    assert '<span class="product-name">INSALATA MISTA - SENZA GLUTINE E SENZA LATTOSIO</span>' in mobile_page.text


def test_admin_can_create_cashier_with_assigned_customer_printer(admin_client, db_session):
    default_customer = db_session.scalar(select(Printer).where(Printer.is_customer_printer.is_(True)))

    printer_response = admin_client.post(
        "/printers",
        data={
            "name": "Cassa 2 Cliente",
            "type": "fake",
            "port": "9100",
            "enabled": "true",
            "is_customer_printer": "true",
        },
        follow_redirects=False,
    )
    assert printer_response.status_code == 303
    second_customer = db_session.scalar(select(Printer).where(Printer.name == "Cassa 2 Cliente"))
    assert second_customer is not None
    db_session.refresh(default_customer)
    assert default_customer.is_customer_printer is True

    user_response = admin_client.post(
        "/users",
        data={
            "name": "cassa2",
            "pin": "2222",
            "role": "cashier",
            "active": "true",
            "customer_printer_id": str(second_customer.id),
        },
        follow_redirects=False,
    )
    assert user_response.status_code == 303
    user = db_session.scalar(select(User).where(User.name == "cassa2"))
    assert user is not None
    assert user.role == "cashier"
    assert user.customer_printer_id == second_customer.id

    users_page = admin_client.get("/users")
    assert users_page.status_code == 200
    assert "cassa2" in users_page.text
    assert "Cassa 2 Cliente" in users_page.text


def test_logged_in_cashier_uses_assigned_customer_printer(client, db_session):
    assigned_printer = Printer(name="Cassa 2 Cliente", type="fake", enabled=True, is_customer_printer=True)
    db_session.add(assigned_printer)
    db_session.flush()
    cashier = User(
        name="cassa2",
        pin_hash=hash_pin("2222"),
        role="cashier",
        active=True,
        customer_printer_id=assigned_printer.id,
    )
    db_session.add(cashier)
    db_session.commit()
    db_session.refresh(assigned_printer)

    login = client.post("/login", data={"pin": "2222"}, follow_redirects=False)
    assert login.status_code == 303
    product = db_session.scalar(select(Product).where(Product.name == "Acqua"))
    response = client.post(
        "/orders",
        data={"cart_json": json.dumps([{"product_id": product.id, "quantity": 1}])},
        follow_redirects=False,
    )
    assert response.status_code == 303

    order = db_session.scalar(select(Order).order_by(Order.id.desc()))
    customer_job = db_session.scalar(
        select(PrintJob).where(PrintJob.order_id == order.id, PrintJob.job_type == "customer")
    )
    assert customer_job is not None
    assert customer_job.printer_id == assigned_printer.id


def test_printer_delete_unassigns_references(admin_client, db_session):
    printer = db_session.scalar(select(Printer).where(Printer.name == "Kitchen Printer"))
    category = db_session.scalar(select(Category).where(Category.printer_id == printer.id))
    product = db_session.scalar(select(Product).where(Product.category_id == category.id))
    order = create_confirmed_order(
        db_session,
        parse_cart_json(json.dumps([{"product_id": product.id, "quantity": 1}])),
        source="test",
    )
    result = printing.print_order(db_session, order.id, include_customer=False, include_production=True)

    printer_id = printer.id
    category_id = category.id
    item_id = order.items[0].id
    job_id = result.jobs[0].id
    assert order.items[0].printer_id == printer_id
    assert result.jobs[0].printer_id == printer_id

    response = admin_client.post(f"/printers/{printer_id}/delete", follow_redirects=False)

    assert response.status_code == 303
    assert response.headers["location"] == "/printers"
    db_session.expire_all()
    assert db_session.get(Printer, printer_id) is None
    assert db_session.get(Category, category_id).printer_id is None
    assert db_session.get(OrderItem, item_id).printer_id is None
    assert db_session.get(PrintJob, job_id).printer_id is None


def test_product_delete_and_bulk_actions(admin_client, db_session):
    cucina = db_session.scalar(select(Category).where(Category.name == "Cucina"))
    product_a = Product(name="Bulk A", price_cents=100, category_id=cucina.id, active=True, sort_order=99)
    product_b = Product(name="Bulk B", price_cents=200, category_id=cucina.id, active=True, sort_order=100)
    db_session.add_all([product_a, product_b])
    db_session.commit()
    db_session.refresh(product_a)
    db_session.refresh(product_b)
    product_a_id = product_a.id
    product_b_id = product_b.id

    deactivate = admin_client.post(
        "/products/bulk",
        data={"action": "deactivate", "product_ids": [str(product_a_id), str(product_b_id)], "return_to": "/products"},
        follow_redirects=False,
    )
    assert deactivate.status_code == 303
    db_session.refresh(product_a)
    db_session.refresh(product_b)
    assert product_a.active is False
    assert product_b.active is False

    activate = admin_client.post(
        "/products/bulk",
        data={"action": "activate", "product_ids": [str(product_a_id), str(product_b_id)], "return_to": "/products"},
        follow_redirects=False,
    )
    assert activate.status_code == 303
    db_session.refresh(product_a)
    db_session.refresh(product_b)
    assert product_a.active is True
    assert product_b.active is True

    order = create_confirmed_order(
        db_session,
        parse_cart_json(json.dumps([{"product_id": product_a_id, "quantity": 1}])),
        source="test",
    )
    delete = admin_client.post(
        f"/products/{product_a_id}/delete",
        data={"return_to": "/products"},
        follow_redirects=False,
    )
    assert delete.status_code == 303
    db_session.expire_all()
    assert db_session.get(Product, product_a_id) is None
    db_session.refresh(order.items[0])
    assert order.items[0].product_id is None
    assert order.items[0].product_name == "Bulk A"

    bulk_delete = admin_client.post(
        "/products/bulk",
        data={"action": "delete", "product_ids": [str(product_b_id)], "return_to": "/products"},
        follow_redirects=False,
    )
    assert bulk_delete.status_code == 303
    db_session.expire_all()
    assert db_session.get(Product, product_b_id) is None


def test_order_creation_numbering_snapshot_and_fake_output(cashier_client, db_session, test_env):
    product = db_session.scalar(select(Product).where(Product.name == "Panino salamella"))
    response = cashier_client.post(
        "/orders",
        data={"cart_json": json.dumps([{"product_id": product.id, "quantity": 2, "notes": "senza cipolla"}])},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == "/"

    order = db_session.scalar(select(Order).order_by(Order.id.desc()))
    assert order is not None
    assert order.order_number == 1
    assert order.business_date == business_date_for()
    assert order.items[0].product_name == "Panino salamella"
    assert order.items[0].unit_price_cents == 600
    assert order.items[0].notes == "senza cipolla"

    product.name = "Nome cambiato"
    product.price_cents = 999
    db_session.commit()
    db_session.refresh(order.items[0])
    assert order.items[0].product_name == "Panino salamella"
    assert order.items[0].unit_price_cents == 600

    jobs = db_session.scalars(select(PrintJob).where(PrintJob.order_id == order.id)).all()
    assert len(jobs) == 2
    assert {job.job_type for job in jobs} == {"customer", "production"}
    assert all(job.status == "printed" for job in jobs)

    output_files = list(Path(test_env["print_output_dir"]).glob("*.txt"))
    assert len(output_files) == 2
    assert any("ORDINE" in path.read_text(encoding="utf-8") and "001" in path.read_text(encoding="utf-8") for path in output_files)
    customer_ticket = next(path.read_text(encoding="utf-8") for path in output_files if "_customer_" in path.name)
    assert "COMANDA CLIENTE" in customer_ticket
    assert " 2 x PANINO SALAMELLA" in customer_ticket
    assert "2 x 6.00 EUR = 12.00 EUR" in customer_ticket
    assert "*" * printing.CUSTOMER_TICKET_WIDTH in customer_ticket
    assert "TOTALE" in customer_ticket
    assert "12.00 EUR" in customer_ticket
    assert all(len(line) <= printing.CUSTOMER_TICKET_WIDTH for line in customer_ticket.splitlines())
    customer_ticket_lines = customer_ticket.split("\n")
    assert customer_ticket_lines[-4].strip() == "Grazie!"
    assert customer_ticket_lines[-3:] == ["", "", ""]
    production_ticket = next(path.read_text(encoding="utf-8") for path in output_files if "_production_" in path.name)
    assert "QTA 2" in production_ticket
    assert "PANINO SALAMELLA" in production_ticket
    assert "NOTE:" in production_ticket
    assert "senza cipolla" in production_ticket


def test_daily_order_number_increment(db_session):
    product = db_session.scalar(select(Product).where(Product.name == "Panino salamella"))
    order_one = create_confirmed_order(db_session, parse_cart_json(json.dumps([{"product_id": product.id, "quantity": 1}])), source="test")
    order_two = create_confirmed_order(db_session, parse_cart_json(json.dumps([{"product_id": product.id, "quantity": 1}])), source="test")
    assert order_one.order_number == 1
    assert order_two.order_number == 2
    assert order_one.business_date == order_two.business_date


def test_grouping_items_by_printer(db_session):
    product_food = db_session.scalar(select(Product).where(Product.name == "Panino salamella"))
    product_bar = db_session.scalar(select(Product).where(Product.name == "Birra media"))
    order = create_confirmed_order(
        db_session,
        parse_cart_json(
            json.dumps(
                [
                    {"product_id": product_food.id, "quantity": 1},
                    {"product_id": product_bar.id, "quantity": 1},
                ]
            )
        ),
        source="test",
    )
    grouped, unassigned = printing.group_items_by_printer(order)
    assert not unassigned
    assert len(grouped) == 2


def test_failed_network_printer_job_is_recorded(db_session, monkeypatch):
    customer_printer = db_session.scalar(select(Printer).where(Printer.is_customer_printer.is_(True)))
    customer_printer.type = "network_escpos"
    customer_printer.ip = "192.0.2.10"
    db_session.commit()
    attempts = []

    def fail_network(job, printer, order):
        attempts.append(job.id)
        raise printing.NetworkPrintError("offline")

    monkeypatch.setattr(printing, "NETWORK_RETRY_DELAYS_SECONDS", (0, 0))
    monkeypatch.setattr(printing, "_send_network_escpos", fail_network)
    product = db_session.scalar(select(Product).where(Product.name == "Panino salamella"))
    order = create_confirmed_order(db_session, parse_cart_json(json.dumps([{"product_id": product.id, "quantity": 1}])), source="test")
    result = printing.print_order(db_session, order.id, include_customer=True, include_production=False)
    assert result.jobs[0].status == "failed"
    assert "offline" in result.jobs[0].error_message
    assert "3 tentativi" in result.jobs[0].error_message
    assert len(attempts) == 3
    assert db_session.get(Order, order.id) is not None


def test_network_printer_retry_can_recover(db_session, monkeypatch):
    customer_printer = db_session.scalar(select(Printer).where(Printer.is_customer_printer.is_(True)))
    customer_printer.type = "network_escpos"
    customer_printer.ip = "192.0.2.10"
    db_session.commit()
    attempts = []

    def flaky_network(job, printer, order):
        attempts.append(job.id)
        if len(attempts) == 1:
            raise printing.NetworkPrintError("temporary timeout")

    monkeypatch.setattr(printing, "NETWORK_RETRY_DELAYS_SECONDS", (0, 0))
    monkeypatch.setattr(printing, "_send_network_escpos", flaky_network)
    product = db_session.scalar(select(Product).where(Product.name == "Panino salamella"))
    order = create_confirmed_order(db_session, parse_cart_json(json.dumps([{"product_id": product.id, "quantity": 1}])), source="test")

    result = printing.print_order(db_session, order.id, include_customer=True, include_production=False)

    assert result.jobs[0].status == "printed"
    assert result.jobs[0].error_message is None
    assert len(attempts) == 2


def test_failed_print_jobs_can_be_filtered_and_retried(admin_client, db_session, monkeypatch):
    customer_printer = db_session.scalar(select(Printer).where(Printer.is_customer_printer.is_(True)))
    customer_printer.type = "network_escpos"
    customer_printer.ip = "192.0.2.10"
    db_session.commit()

    def fail_network(job, printer, order):
        raise printing.NetworkPrintError("temporary timeout")

    monkeypatch.setattr(printing, "NETWORK_RETRY_DELAYS_SECONDS", (0, 0))
    monkeypatch.setattr(printing, "_send_network_escpos", fail_network)
    product = db_session.scalar(select(Product).where(Product.name == "Panino salamella"))
    order = create_confirmed_order(db_session, parse_cart_json(json.dumps([{"product_id": product.id, "quantity": 1}])), source="test")
    printing.print_order(db_session, order.id, include_customer=True, include_production=False)

    detail = admin_client.get(f"/orders/{order.id}")
    assert detail.status_code == 200
    assert "Riprova stampe fallite" in detail.text
    assert "Stampe fallite" in detail.text

    failed_list = admin_client.get(f"/orders?date={order.business_date}&print_status=failed")
    assert failed_list.status_code == 200
    assert "Stampa KO" in failed_list.text
    assert "N. 001" in failed_list.text

    def recover_network(job, printer, order):
        return None

    monkeypatch.setattr(printing, "_send_network_escpos", recover_network)
    retry = admin_client.post(f"/orders/{order.id}/reprint/failed", follow_redirects=False)
    assert retry.status_code == 303

    db_session.expire_all()
    jobs = db_session.scalars(select(PrintJob).where(PrintJob.order_id == order.id)).all()
    assert {job.status for job in jobs} == {"retried", "printed"}

    updated_detail = admin_client.get(f"/orders/{order.id}")
    assert updated_detail.status_code == 200
    assert "Riprova stampe fallite" not in updated_detail.text
    updated_failed_list = admin_client.get(f"/orders?date={order.business_date}&print_status=failed")
    assert "N. 001" not in updated_failed_list.text


def test_usb_printer_job_dispatch_is_recorded(db_session, monkeypatch):
    customer_printer = db_session.scalar(select(Printer).where(Printer.is_customer_printer.is_(True)))
    customer_printer.type = "usb_escpos"
    customer_printer.ip = "04b8:0e15"
    customer_printer.port = 0
    db_session.commit()
    sent = {}

    def fake_usb(job, printer, order):
        sent["job_type"] = job.job_type
        sent["printer_ip"] = printer.ip
        sent["order_id"] = order.id

    monkeypatch.setattr(printing, "_send_usb_escpos", fake_usb)
    product = db_session.scalar(select(Product).where(Product.name == "Panino salamella"))
    order = create_confirmed_order(db_session, parse_cart_json(json.dumps([{"product_id": product.id, "quantity": 1}])), source="test")
    result = printing.print_order(db_session, order.id, include_customer=True, include_production=False)
    assert result.jobs[0].status == "printed"
    assert sent == {"job_type": "customer", "printer_ip": "04b8:0e15", "order_id": order.id}


def test_escpos_payload_contains_the_same_text_as_preview(db_session):
    product = db_session.scalar(select(Product).where(Product.name == "Panino salamella"))
    product.name = "Panino salamella con cipolle caramellate e salsa speciale"
    db_session.commit()

    order = create_confirmed_order(
        db_session,
        parse_cart_json(json.dumps([{"product_id": product.id, "quantity": 1}])),
        source="test",
    )
    customer_printer = db_session.scalar(select(Printer).where(Printer.is_customer_printer.is_(True)))
    customer_text = printing.build_customer_ticket(order)
    customer_job = PrintJob(
        order_id=order.id,
        printer_id=customer_printer.id,
        job_type="customer",
        status="pending",
        payload_text=customer_text,
    )
    customer_data = printing._build_escpos_payload(customer_job, order, customer_printer)
    customer_decoded = customer_data.decode("cp858", errors="ignore")

    assert b"\x1d!\x11" in customer_data
    assert customer_data.endswith(b"\x1dV\x00")
    customer_printer.partial_cut = True
    assert printing._build_escpos_payload(customer_job, order, customer_printer).endswith(b"\x1dV\x01")
    assert all(len(line) <= printing.CUSTOMER_TICKET_WIDTH for line in customer_text.splitlines())
    for line in customer_text.splitlines():
        if line.strip():
            assert line.strip() in customer_decoded

    kitchen_printer = db_session.scalar(select(Printer).where(Printer.name == "Kitchen Printer"))
    production_text = printing.build_production_ticket(order, kitchen_printer, [order.items[0]])
    production_job = PrintJob(
        order_id=order.id,
        printer_id=kitchen_printer.id,
        job_type="production",
        status="pending",
        payload_text=production_text,
    )
    kitchen_printer.partial_cut = True
    production_data = printing._build_escpos_payload(production_job, order, kitchen_printer)
    production_decoded = production_data.decode("cp858", errors="ignore")
    assert production_data.endswith(b"\x1dV\x01")

    for line in production_text.splitlines():
        if line.strip():
            assert line.strip() in production_decoded


def test_reprint_customer_and_production(cashier_client, db_session):
    product = db_session.scalar(select(Product).where(Product.name == "Patatine"))
    cashier_client.post(
        "/orders",
        data={"cart_json": json.dumps([{"product_id": product.id, "quantity": 1}])},
        follow_redirects=False,
    )
    order = db_session.scalar(select(Order).order_by(Order.id.desc()))
    initial_jobs = len(db_session.scalars(select(PrintJob).where(PrintJob.order_id == order.id)).all())

    customer = cashier_client.post(f"/orders/{order.id}/reprint/customer", follow_redirects=False)
    production = cashier_client.post(f"/orders/{order.id}/reprint/production", follow_redirects=False)
    assert customer.status_code == 303
    assert production.status_code == 303

    jobs = db_session.scalars(select(PrintJob).where(PrintJob.order_id == order.id)).all()
    assert len(jobs) == initial_jobs + 2

    detail = cashier_client.get(f"/orders/{order.id}")
    assert detail.status_code == 200
    assert "Stampe" in detail.text


def test_close_register_creates_sales_history_and_closes_orders(cashier_client, db_session):
    product = db_session.scalar(select(Product).where(Product.name == "Acqua"))
    cashier_client.post(
        "/orders",
        data={"cart_json": json.dumps([{"product_id": product.id, "quantity": 3}])},
        follow_redirects=False,
    )
    order = db_session.scalar(select(Order).order_by(Order.id.desc()))
    assert order.status == "confirmed"
    assert order.paid_at is None

    response = cashier_client.post(
        "/closures",
        data={"business_date": order.business_date, "register_session": "1", "notes": "fine serata"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    closure = db_session.scalar(select(RegisterClosure).where(RegisterClosure.business_date == order.business_date))
    assert closure is not None
    assert closure.register_session == 1
    assert closure.order_count == 1
    assert closure.sales_total_cents == 300
    assert closure.cash_total_cents == 300

    db_session.refresh(order)
    assert order.status == "delivered"
    assert order.paid_at is not None
    assert order.completed_at is not None

    detail = cashier_client.get(f"/closures/{closure.id}")
    assert detail.status_code == 200
    assert "Chiusura" in detail.text


def test_order_number_resets_after_register_closure(cashier_client, db_session):
    product = db_session.scalar(select(Product).where(Product.name == "Caffe"))
    first_response = cashier_client.post(
        "/orders",
        data={"cart_json": json.dumps([{"product_id": product.id, "quantity": 1}])},
        follow_redirects=False,
    )
    assert first_response.status_code == 303
    first_order = db_session.scalar(select(Order).order_by(Order.id.desc()))
    assert first_order.order_number == 1
    assert first_order.register_session == 1

    close_response = cashier_client.post(
        "/closures",
        data={"business_date": first_order.business_date, "register_session": "1"},
        follow_redirects=False,
    )
    assert close_response.status_code == 303

    second_response = cashier_client.post(
        "/orders",
        data={"cart_json": json.dumps([{"product_id": product.id, "quantity": 1}])},
        follow_redirects=False,
    )
    assert second_response.status_code == 303
    second_order = db_session.scalar(select(Order).order_by(Order.id.desc()))
    assert second_order.business_date == first_order.business_date
    assert second_order.register_session == 2
    assert second_order.order_number == 1


def test_mobile_order_is_confirmed_printed_and_returns_to_mobile(cashier_client, db_session):
    product = db_session.scalar(select(Product).where(Product.name == "Birra media"))
    response = cashier_client.post(
        "/mobile/orders",
        data={"cart_json": json.dumps([{"product_id": product.id, "quantity": 1}])},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == "/mobile"

    order = db_session.scalar(select(Order).order_by(Order.id.desc()))
    assert order.source == "mobile"
    assert order.status == "confirmed"
    assert order.order_number == 1
    assert db_session.scalars(select(PrintJob).where(PrintJob.order_id == order.id)).all()


def test_reservation_import_upload_from_browser(admin_client, db_session):
    import openpyxl

    page = admin_client.get("/reservations")
    assert page.status_code == 200
    assert 'type="file"' in page.text
    assert 'name="import_path"' not in page.text

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Informazioni cronologiche",
            "Indirizzo email",
            "Cognome ",
            "Nome",
            "Numero di partecipanti",
            "Tipologia di prenotazione",
            "Liberatoria allergeni",
            "€2,50 DOLCE PRENOTATO",
        ]
    )
    sheet.append(["2026-07-01 09:30:00", "bianchi@example.com", "Bianchi", "Anna", 1, "Ospite", "Ok", 2])
    upload = BytesIO()
    workbook.save(upload)
    upload.seek(0)

    response = admin_client.post(
        "/reservations/import",
        files={
            "import_file": (
                "prenotazioni.xlsx",
                upload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == "/reservations"

    db_session.expire_all()
    reservation = db_session.scalar(select(Reservation).where(Reservation.last_name == "Bianchi"))
    assert reservation is not None
    assert reservation.source_file == "prenotazioni.xlsx"
    assert reservation.total_cents == 500


def test_reservation_import_and_checkout_from_cashier(admin_client, db_session, tmp_path):
    import openpyxl

    path = tmp_path / "prenotazioni.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Risposte del modulo 1"
    sheet.append(
        [
            "Informazioni cronologiche",
            "Indirizzo email",
            "Cognome ",
            "Nome",
            "Numero di partecipanti",
            "Tipologia di prenotazione",
            "Liberatoria allergeni",
            "€4,00 PANINO CON SALAMELLA",
            "€3,00 PATATINE FRITTE",
        ]
    )
    sheet.append(
        [
            "2026-07-01 10:00:00",
            "rossi@example.com",
            "Rossi",
            "Mario",
            2,
            "Ospite",
            "Ho letto e compreso",
            2,
            1,
        ]
    )
    workbook.save(path)

    result = import_reservations_from_xlsx(db_session, str(path))
    assert result.created == 1
    assert result.products_created == 2

    reservation = db_session.scalar(select(Reservation).where(Reservation.last_name == "Rossi"))
    assert reservation is not None
    assert reservation.total_cents == 1100
    assert len(reservation.items) == 2
    panino_alias = db_session.scalar(select(ProductImportAlias).where(ProductImportAlias.source_name == "PANINO CON SALAMELLA"))
    assert panino_alias is not None
    panino_product = db_session.get(Product, panino_alias.product_id)
    assert panino_product is not None

    panino_product.name = "Panino salamella pren."
    db_session.commit()
    second_import = import_reservations_from_xlsx(db_session, str(path))
    assert second_import.products_created == 0
    db_session.expire_all()
    reservation = db_session.scalar(select(Reservation).where(Reservation.last_name == "Rossi"))
    assert any(item.product_name == "Panino salamella pren." for item in reservation.items)

    page = admin_client.get("/reservations?q=rossi")
    assert page.status_code == 200
    assert "Rossi Mario" in page.text
    assert "Apri in cassa" in page.text

    checkout_page = admin_client.get(f"/?reservation_id={reservation.id}")
    assert checkout_page.status_code == 200
    assert "Rossi Mario" in checkout_page.text
    assert "window.initialCart" in checkout_page.text
    assert "Panino salamella pren." in checkout_page.text
    assert '<textarea name="notes" rows="2"></textarea>' in checkout_page.text

    create_response = admin_client.post(
        "/orders",
        data={
            "reservation_id": str(reservation.id),
            "cart_json": json.dumps(
                [
                    {"product_id": reservation.items[0].product_id, "quantity": 2},
                    {"product_id": reservation.items[1].product_id, "quantity": 1},
                ]
            ),
            "notes": "prenotazione controllata in cassa",
        },
        follow_redirects=False,
    )
    assert create_response.status_code == 303
    assert create_response.headers["location"] == "/"

    order = db_session.scalar(select(Order).order_by(Order.id.desc()))
    assert order.status == "confirmed"
    assert order.source == "reservation"
    assert order.order_number == 1
    assert order.total_cents == 1100
    assert order.notes == "prenotazione controllata in cassa"
    assert db_session.scalars(select(PrintJob).where(PrintJob.order_id == order.id)).all()
    db_session.refresh(reservation)
    assert reservation.status == "converted"
    assert reservation.order_id == order.id


def test_import_alias_can_be_reassigned_to_existing_cashier_product(admin_client, db_session, tmp_path):
    import openpyxl

    category = db_session.scalar(select(Category).where(Category.name == "Cucina"))
    short_product = Product(name="Menu breve", price_cents=900, category_id=category.id, active=True)
    db_session.add(short_product)
    db_session.commit()
    db_session.refresh(short_product)

    path = tmp_path / "prenotazioni.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Informazioni cronologiche",
            "Indirizzo email",
            "Cognome ",
            "Nome",
            "Numero di partecipanti",
            "Tipologia di prenotazione",
            "Liberatoria allergeni",
            "€9,00 MENU COMPLETO CON DESCRIZIONE MOLTO LUNGA",
        ]
    )
    sheet.append(["2026-07-01 11:00:00", "verdi@example.com", "Verdi", "Luigi", 1, "Ospite", "Ok", 1])
    workbook.save(path)

    import_reservations_from_xlsx(db_session, str(path))
    alias = db_session.scalar(
        select(ProductImportAlias).where(ProductImportAlias.source_name == "MENU COMPLETO CON DESCRIZIONE MOLTO LUNGA")
    )
    assert alias is not None

    products_page = admin_client.get("/products")
    assert products_page.status_code == 200
    assert "Nomi import collegati" in products_page.text
    assert "MENU COMPLETO CON DESCRIZIONE MOLTO LUNGA" in products_page.text

    response = admin_client.post(
        f"/products/import-aliases/{alias.id}",
        data={"product_id": str(short_product.id), "return_to": "/products"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    db_session.expire_all()
    alias = db_session.get(ProductImportAlias, alias.id)
    assert alias.product_id == short_product.id

    import_reservations_from_xlsx(db_session, str(path))
    reservation = db_session.scalar(select(Reservation).where(Reservation.last_name == "Verdi"))
    assert reservation.items[0].product_id == short_product.id
    assert reservation.items[0].product_name == "Menu breve"


def test_reservation_bulk_delete_actions(admin_client, db_session):
    product = db_session.scalar(select(Product).where(Product.name == "Acqua"))
    reservation_a = Reservation(
        source_key="bulk-a",
        source_file="test.xlsx",
        source_row=2,
        first_name="Anna",
        last_name="Aperti",
        participant_count=1,
        status="imported",
        total_cents=product.price_cents,
        items=[
            ReservationItem(
                product_id=product.id,
                product_name=product.name,
                quantity=1,
                unit_price_cents=product.price_cents,
                line_total_cents=product.price_cents,
            )
        ],
    )
    reservation_b = Reservation(
        source_key="bulk-b",
        source_file="test.xlsx",
        source_row=3,
        first_name="Bruno",
        last_name="Aperti",
        participant_count=1,
        status="imported",
        total_cents=product.price_cents,
        items=[
            ReservationItem(
                product_id=product.id,
                product_name=product.name,
                quantity=1,
                unit_price_cents=product.price_cents,
                line_total_cents=product.price_cents,
            )
        ],
    )
    db_session.add_all([reservation_a, reservation_b])
    db_session.commit()
    db_session.refresh(reservation_a)
    db_session.refresh(reservation_b)
    reservation_a_id = reservation_a.id
    reservation_b_id = reservation_b.id

    selected = admin_client.post(
        "/reservations/bulk",
        data={
            "action": "delete_selected",
            "reservation_ids": [str(reservation_a_id)],
            "return_to": "/reservations",
        },
        follow_redirects=False,
    )
    assert selected.status_code == 303
    db_session.expire_all()
    assert db_session.get(Reservation, reservation_a_id) is None
    assert db_session.get(Reservation, reservation_b_id) is not None

    filtered = admin_client.post(
        "/reservations/bulk",
        data={
            "action": "delete_filtered",
            "status": "imported",
            "q": "aperti",
            "return_to": "/reservations?status=imported&q=aperti",
        },
        follow_redirects=False,
    )
    assert filtered.status_code == 303
    db_session.expire_all()
    assert db_session.get(Reservation, reservation_b_id) is None


def test_pending_order_can_be_edited_before_confirm(admin_client, db_session):
    salamella = db_session.scalar(select(Product).where(Product.name == "Panino salamella"))
    patatine = db_session.scalar(select(Product).where(Product.name == "Patatine"))
    order = create_confirmed_order(
        db_session,
        parse_cart_json(json.dumps([{"product_id": salamella.id, "quantity": 1}])),
        source="test",
    )
    order.status = "pending_confirmation"
    order.order_number = None
    db_session.commit()

    response = admin_client.post(
        f"/orders/{order.id}/edit",
        data={
            "cart_json": json.dumps(
                [
                    {"product_id": salamella.id, "quantity": 2, "notes": "ben cotta"},
                    {"product_id": patatine.id, "quantity": 1},
                ]
            ),
            "notes": "modificata prima della stampa",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"] == f"/orders/{order.id}"
    db_session.expire_all()
    edited = db_session.get(Order, order.id)
    assert edited.status == "pending_confirmation"
    assert edited.total_cents == 1550
    assert edited.notes == "modificata prima della stampa"
    assert [(item.product_name, item.quantity, item.notes) for item in edited.items] == [
        ("Panino salamella", 2, "ben cotta"),
        ("Patatine", 1, None),
    ]
    detail = admin_client.get(f"/orders/{order.id}")
    assert detail.status_code == 200
    assert "pending-order-form" in detail.text
    assert "Aggiungi prodotto" in detail.text
