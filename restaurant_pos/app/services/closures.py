from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from typing import Sequence

from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from ..models import Order, RegisterClosure, RegisterClosureProduct, User
from .numbering import current_register_session


class ClosureError(ValueError):
    pass


@dataclass(frozen=True)
class ProductSalesLine:
    category_name: str
    product_name: str
    quantity: int
    sales_total_cents: int


@dataclass(frozen=True)
class SalesSummary:
    business_date: str
    register_session: int
    order_count: int
    sales_total_cents: int
    cash_total_cents: int
    cancelled_count: int
    pending_count: int
    open_count: int
    product_quantity: int
    product_summaries: tuple[ProductSalesLine, ...]
    closure: RegisterClosure | None


SALE_STATUSES = {"confirmed", "paid", "delivered"}


def summarize_products(orders: Sequence[Order]) -> tuple[ProductSalesLine, ...]:
    grouped: dict[tuple[str, str], tuple[int, int]] = {}
    for order in orders:
        for item in order.items:
            key = (item.category_name, item.product_name)
            quantity, total_cents = grouped.get(key, (0, 0))
            grouped[key] = (quantity + item.quantity, total_cents + item.line_total_cents)
    return tuple(
        ProductSalesLine(
            category_name=category_name,
            product_name=product_name,
            quantity=values[0],
            sales_total_cents=values[1],
        )
        for (category_name, product_name), values in sorted(grouped.items())
    )


def get_sales_summary(db: Session, business_date: str, register_session: int | None = None) -> SalesSummary:
    register_session = register_session or current_register_session(db, business_date)
    orders = db.scalars(
        select(Order)
        .where(Order.business_date == business_date, Order.register_session == register_session)
        .options(selectinload(Order.items))
    ).all()
    sales_orders = [order for order in orders if order.status in SALE_STATUSES]
    cash_orders = [order for order in sales_orders if order.paid_at is not None or order.status in {"paid", "delivered"}]
    product_summaries = summarize_products(sales_orders)
    closure = db.scalar(
        select(RegisterClosure)
        .where(
            RegisterClosure.business_date == business_date,
            RegisterClosure.register_session == register_session,
        )
        .options(selectinload(RegisterClosure.product_summaries))
    )
    order_count = len(sales_orders)
    sales_total_cents = sum(order.total_cents for order in sales_orders)
    cash_total_cents = sum(order.total_cents for order in cash_orders)
    cancelled_count = sum(1 for order in orders if order.status == "cancelled")
    if closure is not None:
        order_count = closure.order_count
        sales_total_cents = closure.sales_total_cents
        cash_total_cents = closure.cash_total_cents
        cancelled_count = closure.cancelled_count
        if closure.product_summaries or closure.order_count == 0:
            product_summaries = tuple(
                ProductSalesLine(
                    category_name=product.category_name,
                    product_name=product.product_name,
                    quantity=product.quantity,
                    sales_total_cents=product.sales_total_cents,
                )
                for product in closure.product_summaries
            )
    return SalesSummary(
        business_date=business_date,
        register_session=register_session,
        order_count=order_count,
        sales_total_cents=sales_total_cents,
        cash_total_cents=cash_total_cents,
        cancelled_count=cancelled_count,
        pending_count=sum(1 for order in orders if order.status == "pending_confirmation"),
        open_count=sum(1 for order in orders if order.status in {"confirmed", "paid"}),
        product_quantity=sum(line.quantity for line in product_summaries),
        product_summaries=product_summaries,
        closure=closure,
    )


def close_register(
    db: Session,
    business_date: str,
    user: User,
    notes: str | None = None,
    register_session: int | None = None,
) -> RegisterClosure:
    register_session = register_session or current_register_session(db, business_date)
    if (
        db.scalar(
            select(RegisterClosure).where(
                RegisterClosure.business_date == business_date,
                RegisterClosure.register_session == register_session,
            )
        )
        is not None
    ):
        raise ClosureError("Turno gia chiuso")

    orders = db.scalars(
        select(Order)
        .where(Order.business_date == business_date, Order.register_session == register_session)
        .options(selectinload(Order.items))
    ).all()
    pending_count = sum(1 for order in orders if order.status == "pending_confirmation")
    if pending_count:
        raise ClosureError("Ci sono ordini mobile da confermare o annullare")

    now = datetime.now(UTC)
    sales_orders = [order for order in orders if order.status in SALE_STATUSES]
    product_summaries = summarize_products(sales_orders)
    for order in sales_orders:
        order.status = "delivered"
        order.paid_at = order.paid_at or now
        order.completed_at = order.completed_at or now

    closure = RegisterClosure(
        business_date=business_date,
        register_session=register_session,
        closed_at=now,
        closed_by_user_id=user.id,
        order_count=len(sales_orders),
        sales_total_cents=sum(order.total_cents for order in sales_orders),
        cash_total_cents=sum(order.total_cents for order in sales_orders),
        cancelled_count=sum(1 for order in orders if order.status == "cancelled"),
        notes=notes,
        product_summaries=[
            RegisterClosureProduct(
                category_name=line.category_name,
                product_name=line.product_name,
                quantity=line.quantity,
                sales_total_cents=line.sales_total_cents,
            )
            for line in product_summaries
        ],
    )
    db.add(closure)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise ClosureError("Turno gia chiuso") from exc
    db.refresh(closure)
    return closure


def build_closure_excel(closure: RegisterClosure, products: Sequence[RegisterClosureProduct]) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Riepilogo"
    summary_sheet.append(["CHIUSURA CASSA"])
    summary_sheet["A1"].font = Font(bold=True, size=16)
    summary_sheet.append([])
    summary_sheet.append(["Giornata", closure.business_date])
    summary_sheet.append(["Turno", closure.register_session])
    summary_sheet.append(["Chiusa alle", closure.closed_at.strftime("%Y-%m-%d %H:%M")])
    summary_sheet.append(["Ordini vendita", closure.order_count])
    summary_sheet.append(["Prodotti venduti", sum(product.quantity for product in products)])
    summary_sheet.append(["Vendite", closure.sales_total_cents / 100])
    summary_sheet.append(["Incasso", closure.cash_total_cents / 100])
    summary_sheet.append(["Ordini annullati", closure.cancelled_count])
    if closure.notes:
        summary_sheet.append(["Note", closure.notes])
    summary_sheet["B8"].number_format = '0.00 "EUR"'
    summary_sheet["B9"].number_format = '0.00 "EUR"'
    summary_sheet.column_dimensions["A"].width = 22
    summary_sheet.column_dimensions["B"].width = 30

    products_sheet = workbook.create_sheet("Prodotti")
    products_sheet.append(["Categoria", "Prodotto", "Quantita", "Vendite"])
    header_fill = PatternFill(fill_type="solid", fgColor="176B5B")
    for cell in products_sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    for product in products:
        products_sheet.append(
            [
                product.category_name,
                product.product_name,
                product.quantity,
                product.sales_total_cents / 100,
            ]
        )
    for cell in products_sheet["D"][1:]:
        cell.number_format = '0.00 "EUR"'
    products_sheet.freeze_panes = "A2"
    products_sheet.auto_filter.ref = products_sheet.dimensions
    products_sheet.column_dimensions["A"].width = 22
    products_sheet.column_dimensions["B"].width = 42
    products_sheet.column_dimensions["C"].width = 12
    products_sheet.column_dimensions["D"].width = 16

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
