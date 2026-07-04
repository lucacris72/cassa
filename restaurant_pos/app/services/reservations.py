from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, BinaryIO

from sqlalchemy import delete, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from ..config import resolve_project_path
from ..models import Category, Order, Printer, Product, ProductImportAlias, Reservation, ReservationItem
from ..utils import parse_price_to_cents
from .numbering import begin_immediate_if_sqlite
from .orders import CartLine, OrderError, OrderNumberConflict, stage_confirmed_order


class ReservationImportError(ValueError):
    pass


@dataclass(frozen=True)
class ImportResult:
    created: int
    updated: int
    skipped: int
    products_created: int
    products_updated: int


@dataclass(frozen=True)
class ProductColumn:
    index: int
    name: str
    price_cents: int


PRODUCT_HEADER_RE = re.compile(r"^\s*€\s*([0-9]+(?:[,.][0-9]{1,2})?)\s+(.+?)\s*$")


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _quantity(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _timestamp_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return _clean_text(value) or None


def _source_key(*parts: str | None) -> str:
    raw = "|".join(_clean_text(part).lower() for part in parts)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _column_map(headers: list[Any]) -> dict[str, int]:
    normalized = {_clean_text(header).lower(): index for index, header in enumerate(headers)}
    required = {
        "timestamp": "informazioni cronologiche",
        "email": "indirizzo email",
        "last_name": "cognome",
        "first_name": "nome",
        "participant_count": "numero di partecipanti",
        "booking_type": "tipologia di prenotazione",
    }
    result: dict[str, int] = {}
    for key, label in required.items():
        if label not in normalized:
            raise ReservationImportError(f"Colonna mancante nel file prenotazioni: {label}")
        result[key] = normalized[label]
    result["acknowledgement"] = 6 if len(headers) > 6 else result["booking_type"]
    return result


def _product_columns(headers: list[Any]) -> list[ProductColumn]:
    columns: list[ProductColumn] = []
    for index, header in enumerate(headers):
        match = PRODUCT_HEADER_RE.match(_clean_text(header))
        if not match:
            continue
        price_cents = parse_price_to_cents(match.group(1))
        name = _clean_text(match.group(2))
        columns.append(ProductColumn(index=index, name=name, price_cents=price_cents))
    if not columns:
        raise ReservationImportError("Nessuna colonna prodotto trovata nel file prenotazioni")
    return columns


def _reservation_category(db: Session) -> Category:
    category = db.scalar(select(Category).where(Category.name == "Prenotazioni"))
    if category is not None:
        return category

    kitchen_printer = db.scalar(select(Printer).where(Printer.name == "Kitchen Printer"))
    kitchen_category = db.scalar(select(Category).where(Category.name == "Cucina"))
    printer_id = kitchen_category.printer_id if kitchen_category else kitchen_printer.id if kitchen_printer else None
    category = Category(name="Prenotazioni", printer_id=printer_id, active=True, sort_order=5)
    db.add(category)
    db.flush()
    return category


def _product_for_column(db: Session, column: ProductColumn, category: Category) -> tuple[Product, bool, bool]:
    alias = db.scalar(
        select(ProductImportAlias)
        .where(ProductImportAlias.source_name == column.name)
        .options(selectinload(ProductImportAlias.product))
    )
    product = alias.product if alias is not None else None
    created = False
    updated = False
    if product is None:
        product = db.scalar(select(Product).where(Product.name == column.name))
        if product is None:
            product = Product(
                name=column.name,
                price_cents=column.price_cents,
                category_id=category.id,
                active=True,
                sort_order=category.sort_order,
            )
            db.add(product)
            db.flush()
            created = True
    else:
        if product.price_cents != column.price_cents:
            product.price_cents = column.price_cents
            updated = True
        if not product.active:
            product.active = True
            updated = True
    if product.price_cents != column.price_cents:
        product.price_cents = column.price_cents
        updated = True
    if not product.active:
        product.active = True
        updated = True
    if alias is None:
        db.add(
            ProductImportAlias(
                source_name=column.name,
                product_id=product.id,
                source_price_cents=column.price_cents,
            )
        )
    else:
        alias.source_price_cents = column.price_cents
        alias.updated_at = datetime.now(UTC)
    return product, created, updated


def _validate_import_filename(filename: str) -> str:
    source_name = Path((filename or "").strip()).name
    if not source_name:
        raise ReservationImportError("Seleziona un file prenotazioni da importare")
    if Path(source_name).suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ReservationImportError("Formato non supportato: usa un file .xlsx scaricato da Google Moduli")
    return source_name


def _resolve_import_path(import_path: str) -> Path:
    path = Path(import_path.strip() or "prenotazioni.xlsx")
    _validate_import_filename(path.name)
    resolved = resolve_project_path(path)
    if not resolved.exists():
        raise ReservationImportError(f"File non trovato: {resolved}")
    return resolved


def _import_reservations_from_workbook(db: Session, workbook: Any, source_name: str) -> ImportResult:
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ReservationImportError("Il file prenotazioni non contiene risposte")

    headers = list(rows[0])
    columns = _column_map(headers)
    product_columns = _product_columns(headers)

    category = _reservation_category(db)
    product_by_column: dict[int, Product] = {}
    products_created = 0
    products_updated = 0
    for product_column in product_columns:
        product, created, updated = _product_for_column(db, product_column, category)
        product_by_column[product_column.index] = product
        products_created += int(created)
        products_updated += int(updated)

    created_count = 0
    updated_count = 0
    skipped_count = 0

    for row_number, row in enumerate(rows[1:], start=2):
        timestamp = _timestamp_text(row[columns["timestamp"]] if columns["timestamp"] < len(row) else None)
        email = _clean_text(row[columns["email"]] if columns["email"] < len(row) else None) or None
        first_name = _clean_text(row[columns["first_name"]] if columns["first_name"] < len(row) else None)
        last_name = _clean_text(row[columns["last_name"]] if columns["last_name"] < len(row) else None)
        if not first_name and not last_name:
            skipped_count += 1
            continue

        reservation_items: list[ReservationItem] = []
        total_cents = 0
        for product_column in product_columns:
            if product_column.index >= len(row):
                continue
            quantity = _quantity(row[product_column.index])
            if quantity <= 0:
                continue
            product = product_by_column[product_column.index]
            line_total = product.price_cents * quantity
            total_cents += line_total
            reservation_items.append(
                ReservationItem(
                    product_id=product.id,
                    product_name=product.name,
                    quantity=quantity,
                    unit_price_cents=product.price_cents,
                    line_total_cents=line_total,
                )
            )

        if not reservation_items:
            skipped_count += 1
            continue

        source_key = _source_key(timestamp, email, first_name, last_name)
        reservation = db.scalar(
            select(Reservation)
            .where(Reservation.source_key == source_key)
            .options(selectinload(Reservation.items))
        )
        if reservation is None:
            reservation = Reservation(source_key=source_key, source_file=source_name, source_row=row_number)
            db.add(reservation)
            created_count += 1
        elif reservation.status == "converted":
            skipped_count += 1
            continue
        else:
            db.execute(delete(ReservationItem).where(ReservationItem.reservation_id == reservation.id))
            reservation.items = []
            updated_count += 1

        reservation.source_file = source_name
        reservation.source_row = row_number
        reservation.response_timestamp = timestamp
        reservation.email = email
        reservation.first_name = first_name
        reservation.last_name = last_name
        reservation.participant_count = _quantity(row[columns["participant_count"]] if columns["participant_count"] < len(row) else None)
        reservation.booking_type = _clean_text(row[columns["booking_type"]] if columns["booking_type"] < len(row) else None) or None
        reservation.acknowledgement = _clean_text(row[columns["acknowledgement"]] if columns["acknowledgement"] < len(row) else None) or None
        reservation.status = "imported"
        reservation.total_cents = total_cents
        reservation.items = reservation_items

    db.commit()
    return ImportResult(
        created=created_count,
        updated=updated_count,
        skipped=skipped_count,
        products_created=products_created,
        products_updated=products_updated,
    )


def import_reservations_from_xlsx(db: Session, import_path: str) -> ImportResult:
    try:
        import openpyxl
    except ImportError as exc:
        raise ReservationImportError("openpyxl non installato: esegui pip install -r requirements.txt") from exc

    resolved = _resolve_import_path(import_path)
    workbook = openpyxl.load_workbook(resolved, data_only=True)
    return _import_reservations_from_workbook(db, workbook, resolved.name)


def import_reservations_from_upload(db: Session, file: BinaryIO, filename: str) -> ImportResult:
    try:
        import openpyxl
    except ImportError as exc:
        raise ReservationImportError("openpyxl non installato: esegui pip install -r requirements.txt") from exc

    source_name = _validate_import_filename(filename)
    file.seek(0)
    workbook = openpyxl.load_workbook(file, data_only=True)
    return _import_reservations_from_workbook(db, workbook, source_name)


def search_reservations(db: Session, query: str = "", status: str = "open") -> list[Reservation]:
    statement = select(Reservation).options(selectinload(Reservation.items), selectinload(Reservation.order))
    if status == "open":
        statement = statement.where(Reservation.status == "imported")
    elif status in {"imported", "converted"}:
        statement = statement.where(Reservation.status == status)
    query = query.strip()
    if query:
        like = f"%{query}%"
        statement = statement.where(
            Reservation.first_name.ilike(like)
            | Reservation.last_name.ilike(like)
            | Reservation.email.ilike(like)
        )
    return db.scalars(statement.order_by(Reservation.last_name, Reservation.first_name, Reservation.id)).all()


def _reservation_with_items(db: Session, reservation_id: int) -> Reservation | None:
    return db.scalar(
        select(Reservation)
        .where(Reservation.id == reservation_id)
        .options(
            selectinload(Reservation.items).selectinload(ReservationItem.product),
            selectinload(Reservation.order),
        )
    )


def load_reservation_for_checkout(db: Session, reservation_id: int) -> Reservation:
    reservation = _reservation_with_items(db, reservation_id)
    if reservation is None:
        raise OrderError("Prenotazione non trovata")
    return reservation


def reservation_order_notes(reservation: Reservation) -> str:
    return (
        f"Prenotazione {reservation.last_name} {reservation.first_name}"
        f" - {reservation.email or 'email non indicata'}"
        f" - {reservation.booking_type or 'tipologia non indicata'}"
        f" - partecipanti {reservation.participant_count}"
    )


def reservation_cart_items(reservation: Reservation) -> list[dict[str, object]]:
    cart_items: list[dict[str, object]] = []
    for item in reservation.items:
        if item.product_id is None or item.quantity <= 0:
            continue
        product = item.product
        cart_items.append(
            {
                "product_id": item.product_id,
                "name": product.name if product is not None else item.product_name,
                "price_cents": product.price_cents if product is not None else item.unit_price_cents,
                "quantity": item.quantity,
                "notes": "",
            }
        )
    if not cart_items:
        raise OrderError("Prenotazione senza prodotti ordinabili")
    return cart_items


def create_confirmed_order_from_reservation(
    db: Session,
    reservation_id: int,
    lines: list[CartLine],
    *,
    notes: str | None = None,
    mark_paid: bool = False,
) -> Order:
    try:
        begin_immediate_if_sqlite(db)
        reservation = _reservation_with_items(db, reservation_id)
        if reservation is None:
            raise OrderError("Prenotazione non trovata")
        if reservation.status == "converted" and reservation.order_id is not None:
            raise OrderError("Prenotazione gia convertita in comanda")
        order = stage_confirmed_order(
            db,
            lines,
            source="reservation",
            notes=notes,
            mark_paid=mark_paid,
        )
        reservation.status = "converted"
        reservation.order_id = order.id
        db.commit()
        db.refresh(order)
        return order
    except IntegrityError as exc:
        db.rollback()
        raise OrderNumberConflict("Numero ordine duplicato, riprovare") from exc
    except Exception:
        db.rollback()
        raise


def create_order_from_reservation(db: Session, reservation_id: int) -> Order:
    reservation = db.scalar(
        select(Reservation)
        .where(Reservation.id == reservation_id)
        .options(selectinload(Reservation.items))
    )
    if reservation is None:
        raise OrderError("Prenotazione non trovata")
    if reservation.status == "converted" and reservation.order_id is not None:
        order = db.get(Order, reservation.order_id)
        if order is not None:
            return order
    lines = [
        CartLine(product_id=item.product_id, quantity=item.quantity)
        for item in reservation.items
        if item.product_id is not None and item.quantity > 0
    ]
    if not lines:
        raise OrderError("Prenotazione senza prodotti ordinabili")
    notes = (
        f"Prenotazione {reservation.last_name} {reservation.first_name}"
        f" - {reservation.email or 'email non indicata'}"
        f" - {reservation.booking_type or 'tipologia non indicata'}"
        f" - partecipanti {reservation.participant_count}"
    )
    return create_confirmed_order_from_reservation(db, reservation_id, lines, notes=notes)


def _delete_loaded_reservations(db: Session, reservations: list[Reservation]) -> int:
    for reservation in reservations:
        db.delete(reservation)
    db.commit()
    return len(reservations)


def delete_reservations(db: Session, reservation_ids: list[int]) -> int:
    selected_ids = sorted(set(reservation_ids))
    if not selected_ids:
        return 0
    reservations = db.scalars(
        select(Reservation)
        .where(Reservation.id.in_(selected_ids))
        .options(selectinload(Reservation.items))
    ).all()
    return _delete_loaded_reservations(db, reservations)


def delete_reservations_by_filter(db: Session, *, query: str = "", status: str = "open") -> int:
    reservations = search_reservations(db, query=query, status=status)
    return _delete_loaded_reservations(db, reservations)
